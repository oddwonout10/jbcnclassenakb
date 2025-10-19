"""Generate invitation codes and Supabase-ready CSV from the class roster spreadsheet."""

from __future__ import annotations

import argparse
import csv
import hashlib
import secrets
import string
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency warning
    raise SystemExit(
        "openpyxl is required for this script. Install with `pip install openpyxl`."
    ) from exc


DEFAULT_INPUT = Path("data/Grade 3 JBCN Ena class list.xlsx")
DEFAULT_OUTPUT_DIR = Path("data/generated")


@dataclass
class StudentRow:
    full_name: str
    grade: str
    section: str | None
    mother_name: str | None
    mother_phone: str | None
    father_name: str | None
    father_phone: str | None
    invitation_code: str
    invitation_code_salt: str
    invitation_code_hash: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create invitation codes and Supabase import files from the roster."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to the Excel roster file (.xlsx).",
    )
    parser.add_argument(
        "--grade",
        default="Grade 3",
        help="Grade label to store (used when sheet column is empty).",
    )
    parser.add_argument(
        "--section",
        default="Ena",
        help="Section label to store for each student.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory where CSV outputs will be written.",
    )
    return parser.parse_args()


def normalise_phone(raw: str | None) -> str | None:
    if not raw:
        return None
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    return digits if digits else None


def generate_invitation_code(existing: set[str]) -> str:
    alphabet = string.ascii_uppercase + string.digits
    while True:
        code = "-".join(
            "".join(secrets.choice(alphabet) for _ in range(4)) for _ in range(2)
        )
        if code not in existing:
            existing.add(code)
            return code


def hash_code(salt: str, code: str) -> str:
    digest = hashlib.sha256(f"{salt}{code}".encode("utf-8")).hexdigest()
    return digest


def load_students(path: Path, grade_fallback: str, section: str) -> Iterable[StudentRow]:
    if not path.exists():
        raise FileNotFoundError(f"Roster file not found: {path}")

    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1))
    header_values = [(cell.value or "").strip() for cell in header_cells]

    def find(header: str, start: int = 0) -> int:
        header_lower = header.lower()
        for idx in range(start, len(header_values)):
            if header_values[idx].lower() == header_lower:
                return idx
        raise ValueError(f"Column '{header}' not found in header row: {header_values}")

    required_headers = ["Student Name", "Mother' Name", "Father's Name"]
    missing = []
    for label in required_headers:
        try:
            find(label)
        except ValueError:
            missing.append(label)
    if missing:
        raise ValueError(f"Missing required columns in sheet: {', '.join(missing)}")

    student_idx = find("Student Name")
    mother_idx = find("Mother' Name")
    father_idx = find("Father's Name")
    grade_idx = None
    try:
        grade_idx = find("Grade")
    except ValueError:
        grade_idx = None

    mother_mobile_idx = None
    father_mobile_idx = None
    try:
        mother_mobile_idx = find("Mobile No.", start=mother_idx + 1)
        father_mobile_idx = find("Mobile No.", start=mother_mobile_idx + 1)
    except ValueError:
        # Fallback names for differently labelled columns
        try:
            mother_mobile_idx = find("Mother Mobile No.")
        except ValueError:
            mother_mobile_idx = None
        try:
            father_mobile_idx = find("Father Mobile No.")
        except ValueError:
            father_mobile_idx = None

    codes_seen: set[str] = set()
    results: list[StudentRow] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        student_raw = row[student_idx] if student_idx < len(row) else None
        student_name = (student_raw or "").strip()
        if not student_name:
            continue

        raw_grade = row[grade_idx] if grade_idx is not None and grade_idx < len(row) else None
        grade = str(raw_grade).strip() if raw_grade else grade_fallback

        mother_raw = row[mother_idx] if mother_idx < len(row) else None
        mother_name = (mother_raw or "").strip() or None

        father_raw = row[father_idx] if father_idx < len(row) else None
        father_name = (father_raw or "").strip() or None

        mother_phone = (
            normalise_phone(row[mother_mobile_idx]) if mother_mobile_idx is not None and mother_mobile_idx < len(row) else None
        )
        father_phone = (
            normalise_phone(row[father_mobile_idx]) if father_mobile_idx is not None and father_mobile_idx < len(row) else None
        )

        code = generate_invitation_code(codes_seen)
        salt = secrets.token_hex(8)
        code_hash = hash_code(salt, code)

        results.append(
            StudentRow(
                full_name=student_name,
                grade=grade,
                section=section,
                mother_name=mother_name,
                mother_phone=mother_phone,
                father_name=father_name,
                father_phone=father_phone,
                invitation_code=code,
                invitation_code_salt=salt,
                invitation_code_hash=code_hash,
            )
        )

    return results


def write_supabase_csv(path: Path, rows: Iterable[StudentRow]) -> None:
    fieldnames = [
        "full_name",
        "grade",
        "section",
        "mother_name",
        "mother_phone",
        "father_name",
        "father_phone",
        "invitation_code_salt",
        "invitation_code_hash",
    ]
    with path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "full_name": row.full_name,
                    "grade": row.grade,
                    "section": row.section,
                    "mother_name": row.mother_name or "",
                    "mother_phone": row.mother_phone or "",
                    "father_name": row.father_name or "",
                    "father_phone": row.father_phone or "",
                    "invitation_code_salt": row.invitation_code_salt,
                    "invitation_code_hash": row.invitation_code_hash,
                }
            )


def write_invitation_csv(path: Path, rows: Iterable[StudentRow]) -> None:
    fieldnames = [
        "student_name",
        "grade",
        "section",
        "invitation_code",
        "mother_name",
        "father_name",
    ]
    with path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "student_name": row.full_name,
                    "grade": row.grade,
                    "section": row.section,
                    "invitation_code": row.invitation_code,
                    "mother_name": row.mother_name or "",
                    "father_name": row.father_name or "",
                }
            )


def main() -> None:
    args = parse_args()
    students = load_students(args.input, args.grade, args.section)

    args.output_dir.mkdir(parents=True, exist_ok=True)

    supabase_csv = args.output_dir / "students_supabase.csv"
    invitation_csv = args.output_dir / "invitation_codes.csv"

    write_supabase_csv(supabase_csv, students)
    write_invitation_csv(invitation_csv, students)

    print(f"Wrote {len(students)} student rows for Supabase import -> {supabase_csv}")
    print(f"Wrote invitation code roster for distribution -> {invitation_csv}")


if __name__ == "__main__":
    main()
